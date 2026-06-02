"""
source_lists.py — Load vetted source URLs from your findings file (CSV or XLSX).

Reads the human-validated source inventory you maintain, so the scraper picks
up new sources as you add them — no code change. Columns are matched by HEADER
NAME (case-insensitive, flexible), not position, so you can add/reorder columns
(e.g., an RSS feed column, an organization-name column) freely.

For XLSX, reads the sheet named 'Useable-URL' (the validated tab), leaving your
audit/notes tab untouched.

Recognized headers (only State + a URL are required):
    State          : "state"
    Working URL    : "working url" | "url" | "working_url"
    RSS/Feed URL   : "rss" | "feed" | "feed url" | "rss url"   (optional)
    Organization   : "organization" | "org" | "ag" | "office" | "agency" (optional)

A row is used only if it has a State and at least one usable URL (feed or HTML).
If a feed URL is present, the scraper prefers it over the HTML URL.
"""

import csv
import os


def _norm(s):
    return (s or "").strip().lower()


def _pick(headers, *candidates):
    """Return the index of the first header matching any candidate, else None."""
    norm = [_norm(h) for h in headers]
    for cand in candidates:
        if cand in norm:
            return norm.index(cand)
    return None


def _rows_to_sources(headers, rows):
    i_state = _pick(headers, "state")
    i_url = _pick(headers, "working url", "url", "working_url")
    i_feed = _pick(headers, "rss", "feed", "feed url", "rss url")
    i_org = _pick(headers, "organization", "org", "ag", "office", "agency")

    if i_state is None or i_url is None:
        raise ValueError(
            "Findings file must have at least 'State' and 'Working URL' columns. "
            f"Found headers: {headers}"
        )

    sources = []
    for r in rows:
        def cell(idx):
            return (r[idx].strip() if idx is not None and idx < len(r) and r[idx] else "")
        state = cell(i_state)
        html_url = cell(i_url)
        feed_url = cell(i_feed)
        org = cell(i_org)
        if not state:
            continue
        if not (html_url or feed_url):
            continue  # skip rows with no usable URL yet
        sources.append({
            "state": state,
            "organization": org or f"{state} Attorney General",
            "html_url": html_url or None,
            "feed_url": feed_url or None,
        })
    return sources


def load_from_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        all_rows = [row for row in reader if any(c.strip() for c in row)]
    if not all_rows:
        return []
    return _rows_to_sources(all_rows[0], all_rows[1:])


def load_from_xlsx(path, sheet_name="Useable-URL"):
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError(
            "Reading .xlsx requires openpyxl. Install with: pip install openpyxl"
        ) from e
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found in {path}. Sheets: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c) for c in row]
        if any(c.strip() for c in cells):
            all_rows.append(cells)
    if not all_rows:
        return []
    return _rows_to_sources(all_rows[0], all_rows[1:])


def load_sources(path, sheet_name="Useable-URL"):
    """Load from .csv or .xlsx based on file extension."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return load_from_csv(path)
    if ext in (".xlsx", ".xlsm"):
        return load_from_xlsx(path, sheet_name=sheet_name)
    raise ValueError(f"Unsupported findings file type: {ext} (use .csv or .xlsx)")
